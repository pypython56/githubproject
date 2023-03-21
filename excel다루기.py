import openpyxl

wb=openpyxl.load_workbook('C:/coding/xl.xlsx')
#print(wb.sheetnames) 시트 이름 출력

sheet=wb['Sheet1'] #시트 가지고 오기

# print(sheet['A1'].value)
# print(sheet['b1'].value)

#column (세로) 값 읽기
for cols in sheet.iter_cols():
    for cell in cols:
        print(cell.value,end="\t")
    print()

