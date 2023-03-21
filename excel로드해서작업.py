import openpyxl
from openpyxl.styles import PatternFill, colors

wb=openpyxl.load_workbook("C:/coding/pyexcel.xlsx")
sheet=wb["Test"]

sheet.append([1,2,3,4,5])
sheet.append(["김","이","박","최"])

#배경 색상
pattern_red=PatternFill(start_color="ff0000",fill_type="solid")#RGB값으로 색설정
pattern_blue=PatternFill(start_color=colors.BLUE,fill_type="solid")

sheet.cell(3,1).fill=pattern_red#(3,1)행렬로 생각
sheet.cell(3,3).fill=pattern_blue

wb.save("C:/coding/pyexcel.xlsx")